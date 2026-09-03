"""Экспорт собранных данных в форматированный Excel-отчёт (.xlsx)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = ["Название", "Цена", "Валюта", "Рейтинг (1–5)", "Ссылка"]

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LINK_FONT = Font(color="0563C1", underline="single")


def write_report(rows: list[list], output: str | Path) -> Path:
    """Сформировать .xlsx с оформлением: шапка, фильтры, ширины колонок."""
    path = Path(output)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{max(len(rows) + 1, 2)}"

    for row in rows:
        ws.append(row)

    for idx, width in enumerate((60, 12, 10, 14, 80), start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = "#,##0.00"
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = LINK_FONT

    wb.save(path)
    return path
